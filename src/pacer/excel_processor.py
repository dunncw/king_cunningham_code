import pandas as pd
import re
from difflib import SequenceMatcher
from typing import Dict, List, Tuple, Union, Optional
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

class ExcelFieldMapper:
    COLUMN_MAPPINGS = {
        'account': ['Account #', 'Contract Num'],
        'first_1': ['First 1', 'First Name 1'],
        'middle_1': ['Middle 1', 'Middle Name 1'],
        'last_1': ['Last 1', 'Last Name 1'],
        'first_2': ['First 2', 'First Name 2'],
        'middle_2': ['Middle 2', 'Middle Name 2'],
        'last_2': ['Last 2', 'Last Name 2'],
        'ssn_1': ['SSN 1'],
        'ssn_2': ['SSN 2'],
        'results_1': ['Person_1_Results'],
        'results_2': ['Person_2_Results']
    }
    
    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.mapped_columns = {}
        self._map_columns()
    
    def _map_columns(self):
        for field, possible_names in self.COLUMN_MAPPINGS.items():
            for name in possible_names:
                if name in self.df.columns:
                    self.mapped_columns[field] = name
                    break
    
    def get_column(self, field: str) -> Optional[str]:
        return self.mapped_columns.get(field)
    
    def validate_required_fields(self, use_4digit_mode: bool) -> Tuple[bool, str]:
        required = ['account', 'last_1', 'ssn_1', 'results_1', 'results_2']
        
        if use_4digit_mode:
            required.append('first_1')
        
        missing = []
        for field in required:
            if field not in self.mapped_columns:
                expected_names = ', '.join(self.COLUMN_MAPPINGS[field])
                missing.append(f"{field} (expected: {expected_names})")
        
        if missing:
            return False, f"Missing required columns: {', '.join(missing)}"
        
        return True, ""

class PACERExcelProcessor:
    VALID_RESULTS = ["No Bankruptcy", "Closed Bankruptcy", "OPEN Bankruptcy Found"]
    
    def __init__(self, excel_path, use_4digit_mode=False):
        self.excel_path = excel_path
        self.use_4digit_mode = use_4digit_mode
        self.df = None
        self.field_mapper = None
        self.cells_to_highlight_red = []
        self.cells_to_highlight_yellow = []

    def validate_ssn(self, ssn) -> bool:
        if pd.isna(ssn):
            return False

        ssn_str = str(ssn).split('.')[0]
        ssn_clean = re.sub(r'\D', '', ssn_str)
        
        if self.use_4digit_mode:
            return len(ssn_clean) == 4
        else:
            return len(ssn_clean) == 9

    def format_ssn(self, ssn) -> str:
        ssn_str = str(ssn).split('.')[0]
        ssn_clean = re.sub(r'\D', '', ssn_str)
        return ssn_clean

    def needs_processing(self, result: str) -> bool:
        if pd.isna(result) or not result:
            return True
        result_str = str(result)
        if result_str == 'nan' or not result_str.strip():
            return True
        if result_str not in self.VALID_RESULTS and not result_str.startswith("REVIEW REQUIRED"):
            return True
        return False

    def calculate_name_similarity(self, name1: str, name2: str) -> int:
        if not name1 or not name2:
            return 0
        n1 = name1.strip().upper()
        n2 = name2.strip().upper()
        if n1 == n2:
            return 100
        ratio = SequenceMatcher(None, n1, n2).ratio()
        return int(ratio * 100)

    def process_excel(self) -> Tuple[bool, Union[List[Dict], str]]:
        try:
            self.df = pd.read_excel(
                self.excel_path,
                dtype={
                    'Person_1_Results': str,
                    'Person_2_Results': str
                }
            )
            
            self.field_mapper = ExcelFieldMapper(self.df)
            
            valid, error_msg = self.field_mapper.validate_required_fields(self.use_4digit_mode)
            if not valid:
                return False, error_msg
            
            results_1_col = self.field_mapper.get_column('results_1')
            results_2_col = self.field_mapper.get_column('results_2')
            
            if results_1_col not in self.df.columns:
                self.df[results_1_col] = pd.Series(dtype=str)
            if results_2_col not in self.df.columns:
                self.df[results_2_col] = pd.Series(dtype=str)
            
            self.df[results_1_col] = self.df[results_1_col].astype(str)
            self.df[results_2_col] = self.df[results_2_col].astype(str)
            
            self.df[results_1_col] = self.df[results_1_col].replace('nan', '')
            self.df[results_2_col] = self.df[results_2_col].replace('nan', '')
            
            processed_data = []
            
            for index, row in self.df.iterrows():
                account_col = self.field_mapper.get_column('account')
                row_data = {
                    "excel_row_index": index,
                    "account_number": row[account_col] if pd.notna(row[account_col]) else None,
                    "people": []
                }
                
                last_1_col = self.field_mapper.get_column('last_1')
                ssn_1_col = self.field_mapper.get_column('ssn_1')
                first_1_col = self.field_mapper.get_column('first_1')
                middle_1_col = self.field_mapper.get_column('middle_1')
                
                if (pd.notna(row.get(last_1_col)) and 
                    self.validate_ssn(row.get(ssn_1_col)) and 
                    self.needs_processing(row.get(results_1_col))):
                    person1 = {
                        "person_number": 1,
                        "last_name": row[last_1_col],
                        "ssn": self.format_ssn(row[ssn_1_col]),
                        "first_name": row.get(first_1_col) if pd.notna(row.get(first_1_col)) else None,
                        "middle_name": row.get(middle_1_col) if pd.notna(row.get(middle_1_col)) else None
                    }
                    row_data["people"].append(person1)
                
                last_2_col = self.field_mapper.get_column('last_2')
                ssn_2_col = self.field_mapper.get_column('ssn_2')
                first_2_col = self.field_mapper.get_column('first_2')
                middle_2_col = self.field_mapper.get_column('middle_2')
                
                if (pd.notna(row.get(last_2_col)) and 
                    self.validate_ssn(row.get(ssn_2_col)) and 
                    self.needs_processing(row.get(results_2_col))):
                    person2 = {
                        "person_number": 2,
                        "last_name": row[last_2_col],
                        "ssn": self.format_ssn(row[ssn_2_col]),
                        "first_name": row.get(first_2_col) if pd.notna(row.get(first_2_col)) else None,
                        "middle_name": row.get(middle_2_col) if pd.notna(row.get(middle_2_col)) else None
                    }
                    row_data["people"].append(person2)
                
                if row_data["account_number"] and row_data["people"]:
                    processed_data.append(row_data)
            
            self.save_excel()
            
            return True, processed_data
        except Exception as e:
            return False, f"Error processing Excel file: {str(e)}"

    def update_results(self, row_index: int, person_number: int, result: str, has_exact_match: bool = False) -> Tuple[bool, str]:
        try:
            result_col_name = self.field_mapper.get_column(f'results_{person_number}')
            
            self.df.at[row_index, result_col_name] = result
            
            col_idx = self.df.columns.get_loc(result_col_name) + 1
            row_idx = row_index + 2
            
            if "OPEN Bankruptcy Found" in result:
                self.cells_to_highlight_red.append((col_idx, row_idx))
            elif has_exact_match:
                self.cells_to_highlight_yellow.append((col_idx, row_idx))
            
            self.save_excel()
            
            return True, "Results updated successfully"
            
        except Exception as e:
            print(f"ERROR in update_results: {str(e)}")
            return False, f"Error updating results: {str(e)}"

    def apply_highlighting(self):
        if not self.cells_to_highlight_red and not self.cells_to_highlight_yellow:
            return
            
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            white_font = Font(color="FFFFFF")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            black_font = Font(color="000000")
            
            for col_idx, row_idx in self.cells_to_highlight_red:
                col_letter = get_column_letter(col_idx)
                cell = ws[f"{col_letter}{row_idx}"]
                cell.fill = red_fill
                cell.font = white_font
            
            for col_idx, row_idx in self.cells_to_highlight_yellow:
                col_letter = get_column_letter(col_idx)
                cell = ws[f"{col_letter}{row_idx}"]
                cell.fill = yellow_fill
                cell.font = black_font
            
            wb.save(self.excel_path)
            
            self.cells_to_highlight_red = []
            self.cells_to_highlight_yellow = []
            
        except Exception as e:
            print(f"ERROR applying highlighting: {str(e)}")

    def auto_fit_columns_and_rows(self):
        try:
            from openpyxl.styles import Alignment
            
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            for column_cells in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    try:
                        cell_value = str(cell.value) if cell.value else ""
                        if '\n' in cell_value:
                            lines = cell_value.split('\n')
                            cell_length = max(len(line) for line in lines)
                        else:
                            cell_length = len(cell_value)
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            for row in ws.iter_rows():
                max_lines = 1
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value)
                        line_count = cell_value.count('\n') + 1
                        if line_count > 1:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                        if line_count > max_lines:
                            max_lines = line_count
                if max_lines > 1:
                    ws.row_dimensions[row[0].row].height = max_lines * 15
            
            wb.save(self.excel_path)
            
        except Exception as e:
            print(f"ERROR auto-fitting columns/rows: {str(e)}")

    def save_excel(self) -> bool:
        try:
            self.df.to_excel(self.excel_path, index=False)
            return True
        except Exception as e:
            print(f"Error saving Excel file: {str(e)}")
            return False