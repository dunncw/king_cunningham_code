# test_highlight.py
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def test_highlight(excel_path):
    print(f"Opening workbook: {excel_path}")
    
    # Load the workbook
    wb = load_workbook(excel_path)
    print("Workbook loaded")
    
    # Get the active sheet
    ws = wb.active
    print(f"Active sheet: {ws.title}")
    
    # Get cell A1
    cell = ws['A1']
    print(f"Current value of A1: {cell.value}")
    
    # Create red fill and white font
    red_fill = PatternFill(start_color="FF0000", 
                          end_color="FF0000",
                          fill_type="solid")
    white_font = Font(color="FFFFFF")
    
    # Apply formatting
    cell.fill = red_fill
    cell.font = white_font
    print("Applied red background and white font to A1")
    
    # Save the workbook
    wb.save(excel_path)
    print("Saved workbook")

if __name__ == "__main__":
    # Replace this with your actual Excel file path
    excel_path = r"D:\repositorys\KC_appp\task\pacer_scra\data\in\SSN_11-23-24 copy 2.xlsx"
    test_highlight(excel_path)